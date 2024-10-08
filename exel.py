import openpyxl
# pip install openpyxl
inv_file=openpyxl.load_workbook("inventory.xlsx")
product_list=inv_file["Sheet1"]

# products_per_supplier={"AAA Company":20}
products_per_supplier={}
total_value_per_supplier={}
print(product_list.max_row)

product_under_10_inventory={}

for product_row in range(2,product_list.max_row+1):
    suplier_name=product_list.cell(product_row,4).value
    inventory=product_list.cell(product_row,2).value
    price=product_list.cell(product_row,3).value
    print(suplier_name)
    product_num=product_list.cell(product_row,1).value

    inventory_price=product_list.cell(product_row, 5).value
    # calculation numer of products per supplier
    if suplier_name in products_per_supplier:
        # products_per_supplier["key"]="value"
        current_num_products = products_per_supplier.get(suplier_name)
        products_per_supplier[suplier_name]=current_num_products+1
    else:
        print("adding a new supplier")
        products_per_supplier[suplier_name]=1
    # calculation total value of inventory per supplier
    if suplier_name in total_value_per_supplier:

        current_total_value=total_value_per_supplier.get(suplier_name)
        total_value_per_supplier[suplier_name]=current_total_value+inventory*price
    else:
        total_value_per_supplier[suplier_name]=inventory*price

    # logic products with inventory less than 10
    if inventory<10:
        product_under_10_inventory[int(product_num)]=int(inventory)
    #add value for total inventory price
    inventory_price.value=inventory*price

print(total_value_per_supplier)
print(products_per_supplier)
print(product_under_10_inventory)

inv_file.save("inventory_with_total_value.xlsx")